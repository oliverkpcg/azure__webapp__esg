import streamlit as st
import pandas as pd
import os
import requests
import tempfile
import time
from dotenv import load_dotenv
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

def to_excel():
    response = requests.get("http://localhost:7071/api/azfunc__httptrigger__get_db_information")
    if response.status_code == 200:
        data = response.json()
        dict_of_dfs = {k: pd.DataFrame(v) for k, v in data.items()}
        
    else:
        print(f"Failed to retrieve data. Status code: {response.status_code}")
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    for sheet_name, df in dict_of_dfs.items():
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply formatting
        for col in worksheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light blue background
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin', color='000000'),
                    right=openpyxl.styles.Side(border_style='thin', color='000000'),
                    top=openpyxl.styles.Side(border_style='thin', color='000000'),
                    bottom=openpyxl.styles.Side(border_style='thin', color='000000')
                )

        # Format top row
        for cell in worksheet[1]:
            cell.fill = PatternFill(start_color='A9A9A9', end_color='A9A9A9', fill_type='solid')  # Dark grey background
            cell.font = Font(bold=True)

        # Auto-adjust column width
        dim_holder = DimensionHolder(worksheet=worksheet)
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 2)
            dim_holder[col[0].column_letter] = ColumnDimension(worksheet, min=col[0].column, max=col[0].column, width=adjusted_width)

        worksheet.column_dimensions = dim_holder

    writer.close()  # Correct way to save the Excel file
    output.seek(0)  # Move the cursor to the beginning of the stream
    return output.getvalue()

# Function for Page 1
def page1():
    st.title("Belege automatisch auswerten")
    st.write("Lade ein oder mehrere Anträge hoch. Diese werden ausgewertet und du kannst das Ergebnis als Excel runterladen. In unterer Tabelle kannst du den Status verfolgen.")
    
    # Drag-and-drop file uploader
    uploaded_files = st.file_uploader("Lade die auszuwertenden Files hier hoch", accept_multiple_files=True)

    if st.button('Download Excel'):
        excel_data = to_excel()
        st.download_button(label='Download der aktuellen Daten als Excel.', data=excel_data, file_name='formatted_data.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if uploaded_files:
        with tempfile.TemporaryDirectory() as temp_dir:
            for uploaded_file in uploaded_files:
                file_path = os.path.join(temp_dir, uploaded_file.name)
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                BASE_URL = os.getenv("BASE_URL")
                url = f"{BASE_URL}/api/azfunc__httptrigger__upload_file_to_blob"
                params = {"fp": file_path}
                response = requests.get(url, params=params)
                response.text

            st.write("All files have been uploaded and saved!")

    # Placeholder for the DataFrame table
    st.write("Übersicht ausgewertete Files (Lädt alle 3 Sekunden neu):")
    placeholder = st.empty()

    # Function to simulate fetching/updating data
    def get_data():
        response = requests.get("http://localhost:7071/api/azfunc__httptrigger__get_db_information")
        if response.status_code == 200:
            data = response.json()
            dfs = {k: pd.DataFrame(v) for k, v in data.items()}
            
        else:
            print(f"Failed to retrieve data. Status code: {response.status_code}")

        df = dfs["V__LOGGING"]
        df['ID'] = pd.to_datetime(df['ID'])
        df = df.sort_values(by='ID', ascending=False)
        df['BLOB'] = df['BLOB'].apply(lambda x: '/'.join(x.split('/')[1:]))
        df = df.rename(columns={'BLOB': 'File-Name'})
        df = df.rename(columns={'ID': 'Zeitstempel'})

        return df[["Zeitstempel", "File-Name"]]

    # Loop to refresh data every 3 seconds
    while True:
        df = get_data()  # Fetch or simulate fetching data
        placeholder.table(df)  # Update the table in the placeholder
        time.sleep(3)  # Wait for 3 seconds before refreshing again

# Function for Page 2
def page2():
    st.title("Page 2")
    st.write("Welcome to Page 2!")
    st.write("This is the content of the second page.")

# Main function
def main():
    st.sidebar.title("Navigation")
    page = st.sidebar.radio("Go to", ["Page 1", "Page 2"])

    if page == "Page 1":
        page1()
    elif page == "Page 2":
        page2()

if __name__ == "__main__":
    load_dotenv("/Users/oliverkoehn/repos/aiExamples/azure__funcapp__esg/.env")
    print(f"URL: {os.getenv('BASE_URL')}")
    main()
